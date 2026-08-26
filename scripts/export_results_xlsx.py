"""
Exportação dos resultados do experimento para XLSX (scripts/export_results_xlsx.py)

Lê todos os JSONs de experiment/results/ e gera um arquivo XLSX com 4 abas:
  - Resultados  : uma linha por (modelo × especificação × tratamento)
  - Atividades  : uma linha por atividade individual estimada
  - Pares       : uma linha por par (WH, WD) com métricas derivadas (L_im etc.)
  - Sumário     : contagens e métricas agregadas por modelo

Uso:
    python scripts/export_results_xlsx.py
    python scripts/export_results_xlsx.py --out analysis/resultados.xlsx
"""

import sys
import json
import math
import argparse
from pathlib import Path

ROOT        = Path(__file__).parent.parent
RESULTS_DIR = ROOT / "experiment" / "results"
DEFAULT_OUT = ROOT / "experiment" / "resultados_experimento.xlsx"


# ---------------------------------------------------------------------------
# Leitura e validação
# ---------------------------------------------------------------------------

def iter_results():
    """Itera sobre todos os arquivos JSON em experiment/results/."""
    for model_dir in sorted(RESULTS_DIR.iterdir()):
        if not model_dir.is_dir():
            continue
        for json_file in sorted(model_dir.glob("*.json")):
            with open(json_file) as f:
                yield json.load(f)


def is_valid_estimation(data: dict) -> bool:
    est = data.get("estimation")
    if not isinstance(est, dict):
        return False
    activities = est.get("activities")
    if not isinstance(activities, list) or len(activities) == 0:
        return False
    total = est.get("reported_total")
    if not isinstance(total, (int, float)) or total <= 0:
        return False
    return True


def is_valid_conversion(data: dict) -> bool:
    conv = data.get("conversion")
    if not isinstance(conv, dict):
        return False
    c = conv.get("work_hours_per_workday")
    return isinstance(c, (int, float)) and c > 0


def get_reported_total(data: dict):
    try:
        value = data["estimation"]["reported_total"]["most_likely_effort"]

        if isinstance(value, bool) or not isinstance(value, (int, float)):
            return None

        return float(value)

    except (KeyError, TypeError):
        return None


def get_conversion_factor(data: dict):
    try:
        return float(data["conversion"]["work_hours_per_workday"])
    except (KeyError, TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Construção das linhas por aba
# ---------------------------------------------------------------------------

def build_results_rows(records: list[dict]) -> tuple[list, list]:
    """Retorna (headers, rows) para a aba Resultados."""
    headers = [
        "model_id", "model_used", "spec_id", "treatment",
        "valid_estimation", "valid_conversion",
        "reported_total", "num_activities",
        "conversion_factor_Cim",
        "prompt_tokens", "completion_tokens", "total_tokens",
        "estimation_raw", "conversion_raw",
        "assumptions", "technical_decisions",
    ]
    rows = []
    for r in records:
        est     = r.get("estimation") or {}
        usage   = r.get("usage") or {}
        total   = get_reported_total(r)
        c_im    = get_conversion_factor(r)
        acts    = est.get("activities") or []
        rows.append([
            r.get("model_id", ""),
            r.get("model_used", ""),
            r.get("spec_id", ""),
            r.get("treatment", ""),
            is_valid_estimation(r),
            is_valid_conversion(r) if r.get("treatment") == "WD" else "N/A",
            total,
            len(acts),
            c_im,
            usage.get("prompt_tokens"),
            usage.get("completion_tokens"),
            usage.get("total_tokens"),
            r.get("estimation_raw", ""),
            r.get("conversion_raw", ""),
            "; ".join(est.get("assumptions") or []),
            "; ".join(est.get("technical_and_design_decisions") or []),
        ])
    return headers, rows


def build_activities_rows(records: list[dict]) -> tuple[list, list]:
    """Retorna (headers, rows) para a aba Atividades."""
    headers = [
        "model_id", "spec_id", "treatment",
        "activity_id", "activity_name", "activity_description",
        "most_likely_effort",
    ]
    rows = []
    for r in records:
        est   = r.get("estimation") or {}
        acts  = est.get("activities") or []
        for act in acts:
            rows.append([
                r.get("model_id", ""),
                r.get("spec_id", ""),
                r.get("treatment", ""),
                act.get("activity_id", ""),
                act.get("activity_name", ""),
                act.get("activity_description", ""),
                act.get("most_likely_effort"),
            ])
    return headers, rows


def build_pairs_rows(records: list[dict]) -> tuple[list, list]:
    """
    Retorna (headers, rows) para a aba Pares.

    Cada linha representa um par (WH, WD) para o mesmo (modelo × especificação).
    Calcula Y_WDh = Y_WD × C_im e L_im = ln(Y_WH) - ln(Y_WDh).
    """
    # Indexa por (model_id, spec_id, treatment)
    index: dict[tuple, dict] = {}
    for r in records:
        key = (r.get("model_id", ""), r.get("spec_id", ""), r.get("treatment", ""))
        index[key] = r

    # Coleta pares
    seen_pairs: set[tuple] = set()
    pairs = []
    for (model_id, spec_id, treatment), r in index.items():
        pair_key = (model_id, spec_id)
        if pair_key in seen_pairs:
            continue
        seen_pairs.add(pair_key)

        wh_data = index.get((model_id, spec_id, "WH"))
        wd_data = index.get((model_id, spec_id, "WD"))
        pairs.append((model_id, spec_id, wh_data, wd_data))

    headers = [
        "model_id", "spec_id",
        "wh_valid", "wd_valid",
        "Y_WH", "Y_WD", "C_im",
        "Y_WDh",      # Y_WD × C_im  (WD convertido para work-hours)
        "L_im",       # ln(Y_WH) - ln(Y_WDh)
        "pair_complete",
    ]
    rows = []
    for model_id, spec_id, wh, wd in sorted(pairs):
        wh_valid = wh is not None and is_valid_estimation(wh)
        wd_valid = wd is not None and is_valid_estimation(wd) and is_valid_conversion(wd)

        y_wh  = get_reported_total(wh) if wh else None
        y_wd  = get_reported_total(wd) if wd else None
        c_im  = get_conversion_factor(wd) if wd else None

        y_wdh = None
        l_im  = None
        if y_wd is not None and c_im is not None:
            y_wdh = y_wd * c_im
        if y_wh and y_wdh and y_wh > 0 and y_wdh > 0:
            l_im = math.log(y_wh) - math.log(y_wdh)

        pair_complete = wh_valid and wd_valid and l_im is not None
        rows.append([
            model_id, spec_id,
            wh_valid, wd_valid,
            y_wh, y_wd, c_im,
            y_wdh, l_im,
            pair_complete,
        ])
    return headers, rows


def build_summary_rows(pairs_headers: list, pairs_rows: list) -> tuple[list, list]:
    """Retorna (headers, rows) para a aba Sumário, agrupado por modelo."""
    idx = {h: i for i, h in enumerate(pairs_headers)}

    from collections import defaultdict
    groups: dict[str, list] = defaultdict(list)
    for row in pairs_rows:
        groups[row[idx["model_id"]]].append(row)

    headers = [
        "model_id",
        "total_pairs",
        "complete_pairs",
        "wh_valid",
        "wd_valid",
        "mean_Y_WH",
        "mean_Y_WD",
        "mean_C_im",
        "mean_Y_WDh",
        "mean_L_im",
        "median_L_im",
        "std_L_im",
    ]

    def safe_mean(vals):
        vals = [v for v in vals if v is not None]
        return sum(vals) / len(vals) if vals else None

    def safe_median(vals):
        vals = sorted(v for v in vals if v is not None)
        n = len(vals)
        if n == 0:
            return None
        mid = n // 2
        return vals[mid] if n % 2 else (vals[mid - 1] + vals[mid]) / 2

    def safe_std(vals):
        vals = [v for v in vals if v is not None]
        n = len(vals)
        if n < 2:
            return None
        mu = sum(vals) / n
        return math.sqrt(sum((v - mu) ** 2 for v in vals) / (n - 1))

    rows = []
    for model_id, model_rows in sorted(groups.items()):
        rows.append([
            model_id,
            len(model_rows),
            sum(1 for r in model_rows if r[idx["pair_complete"]]),
            sum(1 for r in model_rows if r[idx["wh_valid"]]),
            sum(1 for r in model_rows if r[idx["wd_valid"]]),
            safe_mean([r[idx["Y_WH"]]  for r in model_rows]),
            safe_mean([r[idx["Y_WD"]]  for r in model_rows]),
            safe_mean([r[idx["C_im"]]  for r in model_rows]),
            safe_mean([r[idx["Y_WDh"]] for r in model_rows]),
            safe_mean([r[idx["L_im"]]  for r in model_rows]),
            safe_median([r[idx["L_im"]] for r in model_rows]),
            safe_std([r[idx["L_im"]]   for r in model_rows]),
        ])
    return headers, rows


# ---------------------------------------------------------------------------
# Escrita XLSX
# ---------------------------------------------------------------------------

def write_xlsx(out_path: Path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl não encontrado. Instale com: pip install openpyxl")
        sys.exit(1)

    records = list(iter_results())
    if not records:
        print(f"Nenhum resultado encontrado em {RESULTS_DIR}")
        sys.exit(0)

    print(f"{len(records)} registros encontrados.")

    res_headers,  res_rows  = build_results_rows(records)
    act_headers,  act_rows  = build_activities_rows(records)
    pair_headers, pair_rows = build_pairs_rows(records)
    sum_headers,  sum_rows  = build_summary_rows(pair_headers, pair_rows)

    wb = Workbook()
    wb.remove(wb.active)  # remove aba padrão vazia

    HEADER_FILL = PatternFill("solid", fgColor="2E4057")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    ALT_FILL    = PatternFill("solid", fgColor="F0F4F8")

    def write_sheet(title: str, headers: list, rows: list):
        ws = wb.create_sheet(title=title)

        # Cabeçalho
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Dados
        for row_i, row in enumerate(rows, start=2):
            fill = ALT_FILL if row_i % 2 == 0 else None
            for col, val in enumerate(row, start=1):
                cell = ws.cell(row=row_i, column=col, value=val)
                if fill:
                    cell.fill = fill
                # Textos longos: wrap
                if isinstance(val, str) and len(val) > 80:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Largura automática
        for col in range(1, len(headers) + 1):
            letter = get_column_letter(col)
            max_len = len(str(headers[col - 1]))
            for row_i in range(2, min(len(rows) + 2, 100)):
                val = ws.cell(row=row_i, column=col).value
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[letter].width = min(max_len + 2, 60)

        ws.freeze_panes = "A2"
        return ws

    write_sheet("Resultados",  res_headers,  res_rows)
    write_sheet("Atividades",  act_headers,  act_rows)
    write_sheet("Pares",       pair_headers, pair_rows)
    write_sheet("Sumário",     sum_headers,  sum_rows)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"XLSX salvo em: {out_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Exporta resultados do experimento para XLSX")
    parser.add_argument(
        "--out", type=Path, default=DEFAULT_OUT,
        help=f"Caminho de saída (padrão: {DEFAULT_OUT.relative_to(ROOT)})"
    )
    args = parser.parse_args()
    write_xlsx(args.out)


if __name__ == "__main__":
    main()
